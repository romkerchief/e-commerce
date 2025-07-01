from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse # Import JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import EmailMessage, send_mail
from django.conf import settings
from django.template.loader import render_to_string, get_template
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import logout, login, authenticate
from django.contrib.auth.forms import UserCreationForm
from django.views import generic
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone
from django.db import transaction
from django.db.models import Count, Sum, F
from django.db.models.functions import TruncMonth
import secrets
from functools import wraps
import smtplib
from .models import Profile, UserLevel, ValidStatus, SellerProfile, Division # type: ignore
from .forms import FormSignUp,FormLogIn, ProfileUpdateForm, SellerSignUpForm, SellerProfileUpdateForm, StaffSignUpForm, OrderStatusUpdateForm, StaffUserBaseForm, StaffProfileEditForm # type: ignore
from home.models import Product, OrderStatus, Category, Order, OrderItem
from xhtml2pdf import pisa #type: ignore
import openpyxl #type: ignore
from openpyxl.writer.excel import save_virtual_workbook #type: ignore

# --- Decorators for Permission Checks ---

def admin_or_hr_required(function):
    """
    Decorator to restrict access to views to Superusers or users in the "SDM" (HR) division.
    """
    @wraps(function)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Please log in to access this page.") # type: ignore
            return redirect('customers:customer_login')

        if request.user.is_superuser:
            return function(request, *args, **kwargs)
        
        try:
            profile = request.user.profile
            if profile.division and profile.division.name == "SDM":
                return function(request, *args, **kwargs)
            else:
                messages.error(request, "You do not have permission to access this page.")
                return redirect('homey:index') # Redirect to homepage or a 'permission denied' page
        except Profile.DoesNotExist:
            messages.error(request, "User profile not found. Please complete your profile or contact support.")
            return redirect('customers:edit_profile')
        except AttributeError: # Handles if profile exists but division is None
            messages.error(request, "User profile is incomplete (division not set). You do not have permission to access this page.")
            return redirect('homey:index')
    return wrap

def staff_member_required(function):
    """
    Decorator to restrict access to views to users with the Django `is_staff` flag set to True.
    Redirects to the customer login page if not authenticated.
    """
    @wraps(function)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Please log in to access this page.")
            return redirect('customers:customer_login') # Use your named URL for login

        if request.user.is_staff: # Check the Django is_staff flag
            return function(request, *args, **kwargs)
        else:
            messages.error(request, "You do not have permission to access this page.")
            return redirect('homey:index') # Redirect non-staff to homepage
    return wrap

def pimpinan_required(function):
    """
    Decorator to restrict access to views to Superusers or users in the "Pimpinan" division.
    """
    @wraps(function)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_staff:
            messages.error(request, "You must be logged in as staff to access this page.")
            return redirect('customers:customer_login')

        if request.user.is_superuser:
            return function(request, *args, **kwargs)
        
        try:
            profile = request.user.profile
            # Check if the user is in the 'Pimpinan' division
            if profile.division and profile.division.name == "Pimpinan":
                return function(request, *args, **kwargs)
            else:
                messages.error(request, "You do not have permission to access the leadership dashboard.")
                return redirect('customers:staff_dashboard') # Redirect to the regular staff dashboard
        except Profile.DoesNotExist:
            messages.error(request, "User profile not found.")
            return redirect('customers:staff_dashboard')
        except AttributeError:
            messages.error(request, "User profile is incomplete (division not set).")
            return redirect('customers:staff_dashboard')
    return wrap
    

# --- Authentication and Profile Management Views ---

@csrf_protect
def registerPage(request):
    """
    Handles user registration. Creates a new user and profile, sends a verification email,
    and logs the user in, redirecting to the email verification page.
    """
    # inisialisasi form (kosong untuk request GET, terisi untuk POST)
    form = FormSignUp()
    
    if request.method == 'POST':
        # hubungkan form dengan data POST
        form = FormSignUp(request.POST)
        
        if form.is_valid():
            try:
                # memulai transaksi atomik untuk memastikan konsistensi data
                with transaction.atomic():
                    print("Form has been posted and is valid")
                    
                    user = form.save()
                    
                    # Profile is now created by the signal in signals.py.
                    # We just need to fetch it to add verification details.
                    # The signal should have already set default user_level and valid_status.
                    profile = user.profile # Access profile created by the signal

                    # Populate profile with additional data from the form
                    profile.full_name = form.cleaned_data.get('full_name')
                    profile.alamat = form.cleaned_data.get('alamat')
                    profile.phone_num = form.cleaned_data.get('phone_num')
                    
                    # Add verification code details
                    verification_code = f"{secrets.randbelow(10**6):06d}"
                    profile.verification_code = verification_code
                    profile.code_created_at = timezone.now()
                    profile.save() # Save all profile changes (full_name, alamat, phone_num, verification details)
                    
                    # mengirim email verifikasi
                    send_mail(
                        'Verify Your Email',
                        f'Your verification code: {verification_code}',
                        'noreply@yourdomain.com',
                        [user.email],
                        fail_silently=False,
                    )
                    
                    # login user dan redirect ke verify.html
                    login(request, user)
                    return redirect('customers:verify_email')
                    
            except smtplib.SMTPException as e:
                # Handle SMTP email sending errors
                error_message = "Gagal mengirimkan Email verifikasi."
                if "550" in str(e):
                    error_message += " Email tidak valid atau tidak ditemukan."
                else:
                    error_message += " Coba lagi nanti, atau hubungi admin."
                
                messages.error(request, error_message)
                # menghapus user jika email gagal dikirim
                if user:
                    user.delete()
                return redirect('customers:register')
                
            except Exception as e:
                # handle error lain
                messages.error(request, "Terjadi kesalahan sistem. Silahkan coba lagi.")
                print(f"Unexpected error during registration: {str(e)}")
                # menghapus user jika ada error
                if 'user' in locals() and user:
                    user.delete()
                return redirect('customers:register')
                
        else:
            # form tidak valid, kirim error message
            print("Form is not valid")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    else:
        print("Loading registration page (GET request)")
    
    # menyiapkan context dan render template
    context = {'form': form}
    return render(request, 'registration/register.html', context)

@csrf_protect
def seller_register_page(request):
    """
    Handles seller registration. Creates a new user and profile with 'Penjual' user_level,
    creates a SellerProfile, sends a verification email, and logs the user in.
    """
    form = SellerSignUpForm()
    if request.method == 'POST':
        form = SellerSignUpForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = form.save() # UserCreationForm handles user creation

                    # Fetch or create the profile (signal should handle creation)
                    profile = user.profile

                    # Set UserLevel to "Penjual"
                    seller_level, created = UserLevel.objects.get_or_create(name="Penjual")
                    profile.user_level = seller_level

                    # Populate other profile fields if collected in SellerSignUpForm
                    profile.full_name = form.cleaned_data.get('full_name')
                    profile.alamat = form.cleaned_data.get('alamat') 
                    profile.phone_num = form.cleaned_data.get('phone_num') 

                    # Add verification code details
                    verification_code = f"{secrets.randbelow(10**6):06d}"
                    profile.verification_code = verification_code
                    profile.code_created_at = timezone.now()
                    profile.save()

                    # The signal for Profile should have created SellerProfile.
                    # Now, populate store_name in SellerProfile.
                    if hasattr(profile, 'seller_specific_profile'):
                        profile.seller_specific_profile.store_name = form.cleaned_data.get('store_name')
                        profile.seller_specific_profile.save()

                    send_mail(
                        'Verify Your Seller Account Email',
                        f'Your verification code: {verification_code}',
                        'noreply@yourdomain.com',
                        [user.email],
                        fail_silently=False,
                    )
                    
                    login(request, user)
                    messages.info(request, "Seller account created. Please verify your email.")
                    return redirect('customers:verify_email') # Or a seller-specific welcome page

            except smtplib.SMTPException as e:
                error_message = "Gagal mengirimkan Email verifikasi."
                if "550" in str(e): error_message += " Email tidak valid atau tidak ditemukan."
                else: error_message += " Coba lagi nanti, atau hubungi admin."
                messages.error(request, error_message)
                if 'user' in locals() and user: user.delete()
                return redirect('customers:seller_register')
            except Exception as e:
                messages.error(request, "Terjadi kesalahan sistem. Silahkan coba lagi.")
                print(f"Unexpected error during seller registration: {str(e)}")
                if 'user' in locals() and user: user.delete()
                return redirect('customers:seller_register')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    
    context = {'form': form}
    return render(request, 'registration/seller_register.html', context)

def loginPage(request):
    form = FormLogIn()
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # First, check if the user is a staff member
            if user.is_staff:
                # If they are staff, perform the status check
                if hasattr(user, 'profile') and user.profile.valid_status:
                    status_name = user.profile.valid_status.name
                    if status_name != 'Valid': # Block if status is anything other than "Valid"
                        messages.error(request, f'This account is not active ({status_name}). Please contact HR.')
                        return render(request, 'registration/login.html', {'form': form})
                else: # Failsafe for misconfigured staff accounts
                    messages.error(request, "Your staff account is not configured correctly. Please contact support.")
                    return render(request, 'registration/login.html', {'form': form})
            
            # If we get here, the user is either not staff, or is a staff member with "Valid" status.
            # It is now safe to log them in.
            login(request, user)

            # Your redirect logic remains the same
            if user.is_staff:
                return redirect('customers:staff_dashboard')
            elif hasattr(user, 'profile') and user.profile.user_level and user.profile.user_level.name == "Penjual":
                return redirect('customers:seller_dashboard')
            else:
                return redirect('homey:index')
        else:
            messages.error(request, 'Invalid credentials')
    
    context = {'form':form}
    return render(request,'registration/login.html', context)

@login_required
def logoutPage(request):
	logout(request)
	return redirect('customers:customer_login')

@login_required
def verify_email(request):
    """
    Handles email verification using a code sent to the user's email.
    Allows the user to enter the code to verify their email address.
    """
    try:
        profile = request.user.profile # Changed to user.profile
    except Profile.DoesNotExist: # Changed to Profile.DoesNotExist
        messages.error(request, "Profil tidak ditemukan. Silahkan daftar ulang.")
        return redirect('customers:register')
    
    if profile.email_verified:
        return redirect('homey:index')

    if request.method == 'POST':
        entered_code = request.POST.get('code', '').strip()
        if entered_code == profile.verification_code:
            if (timezone.now() - profile.code_created_at).total_seconds() <= 86400:
                profile.email_verified = True
                profile.save()
                messages.success(request, 'Email verified!')
                return redirect('homey:index')
            else:
                messages.error(request, "Kode sudah kedaluarsa. Silahkan minta resend")
                return redirect('customers:resend_verification')
        else:
            messages.error(request, 'Invalid code')
    return render(request, 'registration/verify.html')

@login_required
def resend_verification(request):
    """
    Generates and sends a new email verification code to the user's email address.
    Updates the profile with the new code and timestamp.
    """
    profile = request.user.profile # Changed to user.profile
    new_code = f"{secrets.randbelow(10**6):06d}"
    profile.verification_code = new_code
    profile.code_created_at = timezone.now()
    profile.save()
    send_mail(
        'New Verification Code',
        f'Your new code: {new_code}',
        'noreply@example.com',
        [request.user.email],
        fail_silently=False,
    )
    messages.info(request, 'Kode baru telah dikirim ke email Anda')
    return redirect('customers:verify_email')

@login_required
def edit_profile(request):
    """
    Handles editing the user's profile information.
    Includes fields from the main Profile and optionally SellerProfile if the user is a seller.
    Handles email changes by requiring re-verification.
    """
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        messages.error(request, "Profil tidak ditemukan. Silahkan hubungi admin.")
        return redirect('homey:index')

    seller_profile = None
    if hasattr(profile, 'seller_specific_profile'):
        seller_profile = profile.seller_specific_profile

    if request.method == 'POST':
        # 'form_type' will help distinguish which form was submitted if you have multiple submit buttons
        # For simplicity, we'll assume one submit button updates all.
        profile_form = ProfileUpdateForm(request.POST, instance=profile, prefix="profile")
        
        seller_form = None
        if seller_profile:
            seller_form = SellerProfileUpdateForm(request.POST, instance=seller_profile, prefix="seller")

        if profile_form.is_valid() and (not seller_form or seller_form.is_valid()):
            # Save profile fields
            profile_form.save()

            new_email = profile_form.cleaned_data.get('email')
            email_changed = request.user.email != new_email

            if email_changed:
                request.user.email = new_email
                request.user.save()
                
                # Mark email as unverified and send new verification code
                profile.email_verified = False
                verification_code = f"{secrets.randbelow(10**6):06d}"
                profile.verification_code = verification_code
                profile.code_created_at = timezone.now()
                # profile.save() is called below if seller_form is also valid or not present

            if seller_form and seller_form.is_valid():
                seller_form.save()
            
            profile.save() # Save profile changes (including potential verification updates)

            if email_changed:
                send_mail('Verify Your New Email Address', f'Your new verification code: {verification_code}', 'noreply@yourdomain.com', [new_email], fail_silently=False)
                messages.info(request, 'Profile updated. Please verify your new email address.')
                return redirect('customers:verify_email')
            else:
                messages.success(request, 'Profil Anda berhasil diperbarui.')
                return redirect('customers:edit_profile')
        else:
            messages.error(request, "Harap perbaiki kesalahan di bawah ini.")
    else:
        profile_form = ProfileUpdateForm(instance=profile, prefix="profile")
        seller_form = None
        if seller_profile:
            seller_form = SellerProfileUpdateForm(instance=seller_profile, prefix="seller")

    context = {
        'profile_form': profile_form,
        'seller_form': seller_form, # Will be None if not a seller or no seller_profile
    }
    return render(request, 'registration/edit_profile.html', context)

@login_required
def view_profile(request):
    """
    Displays the current user's profile information.
    Fetches the user's Profile object.
    """
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        # This case should ideally not happen if signals are working
        # and user is logged in.
        messages.error(request, "Profil tidak ditemukan. Silahkan hubungi admin.")
        return redirect('homey:index') # Or some other appropriate page
    
    context = {
        'profile': profile,
        'user_object': request.user # Pass the whole user object if you need username, email etc. directly
    }
    return render(request, 'registration/view_profile.html', context)


# --- Seller Dashboard ---

def seller_required(function):
    @wraps(function)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('customers:customer_login')
        try:
            profile = request.user.profile
            if profile.user_level and profile.user_level.name == "Penjual":
                return function(request, *args, **kwargs)
            else:
                messages.error(request, "You do not have permission to access this page.")
                return redirect('homey:index')
        except Profile.DoesNotExist:
            messages.error(request, "User profile not found. Please complete your profile.")
            return redirect('customers:edit_profile')
        except AttributeError: # Handles if profile exists but user_level is None or no seller_specific_profile
            messages.error(request, "User profile incomplete or not configured for seller access.")
            return redirect('homey:index')
    return wrap

@login_required
@admin_or_hr_required
def staff_register_view(request):
    """
    Handles registration of new staff members.
    Restricted to Superusers and users in the "SDM" division.
    """
    if request.method == 'POST':
        form = StaffSignUpForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = form.save(commit=False) # Get unsaved user from UserCreationForm
                    user.is_staff = True  # Mark as Django staff user
                    user.save() # Save user, signal creates/updates profile

                    profile = user.profile # Fetch the profile
                    # Common fields like full_name, alamat, phone_num are in FormSignUp
                    # and should be handled by its fields if they are part of the form.
                    # Explicitly set them if they are not automatically handled by form.save() on profile.
                    profile.full_name = form.cleaned_data.get('full_name')
                    profile.alamat = form.cleaned_data.get('alamat')
                    profile.phone_num = form.cleaned_data.get('phone_num')
                    profile.user_level = form.cleaned_data.get('user_level')
                    profile.division = form.cleaned_data.get('division')

                    profile.email_verified = True # Auto-verify staff email
                    # No verification_code or code_created_at needed for auto-verified staff

                    profile.save() # Save all profile changes

                    send_mail(
                        'Your Staff Account Has Been Created',
                        f'Hello {user.username},\n\nYour staff account has been successfully created. You can now log in with the credentials provided during registration.\n\nThank you.',
                        'noreply@yourdomain.com', # Replace with your actual sending email
                        [user.email],
                        fail_silently=False,
                    )
                    messages.success(request, f"Staff account for {user.username} created and email automatically verified.")
                    return redirect('customers:staff_register') # Or a staff dashboard

            except smtplib.SMTPException as e:
                error_message = "Gagal mengirimkan Email verifikasi."
                if "550" in str(e): error_message += " Email tidak valid atau tidak ditemukan."
                else: error_message += " Coba lagi nanti, atau hubungi admin."
                messages.error(request, error_message)
                if 'user' in locals() and hasattr(user, 'pk') and user.pk: user.delete() # Rollback
            except Exception as e:
                messages.error(request, f"Terjadi kesalahan sistem: {str(e)}")
                print(f"Unexpected error during staff registration: {str(e)}")
                if 'user' in locals() and hasattr(user, 'pk') and user.pk: user.delete() # Rollback
        else:
            for field, errors_list in form.errors.items():
                for error_msg in errors_list:
                    field_label = form.fields[field].label if field in form.fields and hasattr(form.fields[field], 'label') else field.capitalize()
                    messages.error(request, f"{field_label}: {error_msg}")
    else:
        form = StaffSignUpForm()
    
    context = {'form': form, 'form_title': 'Register New Staff Member'}
    return render(request, 'registration/staff_register.html', context)

@login_required
@staff_member_required # Protect this dashboard
def staff_dashboard_view(request):
    """
    Displays the staff dashboard.
    Shows links to different staff functionalities based on the user's division and superuser status.
    """
    profile = request.user.profile if hasattr(request.user, 'profile') else None
    
    # Define division lists for template conditions
    orders_view_divisions = ["SDM", "Penjualan", "Keuangan"]
    products_manage_divisions = ["Penjualan", "Pembelian", "SDM"]
    sdm_division_name = "SDM" # For single division checks

    context = {
        'greeting': f"Welcome to the Staff Dashboard, {request.user.username}!",
        'profile': profile,
        'orders_view_divisions': orders_view_divisions,
        'products_manage_divisions': products_manage_divisions,
        'sdm_division_name': sdm_division_name,
    }
    return render(request, 'staff/staff_dashboard.html', context)

@login_required
@staff_member_required
def staff_all_orders_view(request):
    """
    Staff view to list all orders placed in the system. Includes pagination.
    """
    from home.models import Order # Local import
    # Fetch all orders, potentially filter or sort as needed
    # Show all orders, most recent first
    order_list = Order.objects.all().order_by('-date_order')

    paginator = Paginator(order_list, 15) # Show 15 orders per page
    page_number = request.GET.get('page')
    try:
        orders = paginator.page(page_number)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    context = {
        'orders': orders,
    }
    return render(request, 'staff/all_orders.html', context)

@login_required
@staff_member_required
def staff_order_detail_view(request, order_transaction_id):
    """
    Staff view to display details of a specific order and update its status.
    """
    from home.models import Order, OrderItem, ShippingAddress # type: ignore

    order = get_object_or_404(Order, transaction_id=order_transaction_id)
    order_items = OrderItem.objects.filter(order=order)
    shipping_address = ShippingAddress.objects.filter(order=order).first()

    if request.method == 'POST':
        status_form = OrderStatusUpdateForm(request.POST, instance=order)
        if status_form.is_valid():
            status_form.save()
            messages.success(request, f"Order '{order.transaction_id}' status has been updated successfully.")
            return redirect('customers:staff_order_detail', order_transaction_id=order.transaction_id)
        else:
            messages.error(request, "Failed to update order status. Please review the errors below.")
    else:
        status_form = OrderStatusUpdateForm(instance=order)

    try:
        context = {
            'order': order,
            'order_items': order_items,
            'shipping_address': shipping_address,
            'status_form': status_form,
        }
        return render(request, 'staff/staff_order_detail.html', context)
    except Order.DoesNotExist:
        messages.error(request, "Order not found.")
        return redirect('customers:staff_all_orders')
    except Exception as e:
        messages.error(request, "An error occurred while fetching order details.")
        print(f"Error in staff_order_detail_view: {str(e)}")
        return redirect('customers:staff_all_orders')

@login_required
@seller_required
def seller_dashboard_view(request):
    """
    Displays the seller dashboard. Shows the seller's store name, products, and recent sales.
    """
    seller_profile_obj = None
    products_list = []
    try:
        # print(f"DEBUG: User: {request.user.username}, User Level: {request.user.profile.user_level.name if hasattr(request.user, 'profile') and request.user.profile.user_level else 'N/A'}")
        profile = request.user.profile
        # print(f"[Seller Dashboard] Current user: {request.user.username} (ID: {request.user.id})")

        if hasattr(profile, 'seller_specific_profile'):
            seller_profile_obj = profile.seller_specific_profile
        
        # Check products for this seller
        products_list = Product.objects.filter(seller=request.user).order_by('-upload_time')
        # print(f"[Seller Dashboard] Found {products_list.count()} products for this seller.") # DEBUG
        # for p in products_list: # DEBUG - uncomment to see individual products
        # print(f"[Seller Dashboard] Product: {p.name}, Seller on Product: {p.seller.username if p.seller else 'None'}")

        # Fetch order items for this seller
        from home.models import OrderItem # Import OrderItem
        seller_order_items = OrderItem.objects.filter(product__seller=request.user).select_related('order', 'product', 'order__user').order_by('-order__date_order')


    except Exception as e:
        messages.error(request, "An error occurred while loading your dashboard.")
        print(f"Error in seller_dashboard_view: {str(e)}") # For server logs
        seller_order_items = [] # Ensure it's defined in case of error

    context = {
        'seller_profile': seller_profile_obj, 
        'products': products_list,       # Passed to template as products
        'seller_order_items': seller_order_items,
    }
    return render(request, 'sellers/seller_dashboard.html', context)

@login_required
@seller_required
def seller_order_detail_view(request, order_transaction_id):
    """
    Seller view to display details of a specific order, filtered to show only items from their store.
    """
    try:
        # Fetch the specific order
        # Assuming transaction_id is unique for orders
        from home.models import Order, OrderItem, ShippingAddress
        order = get_object_or_404(Order, transaction_id=order_transaction_id)

        # Filter order items to show only those belonging to this seller
        order_items_for_seller = OrderItem.objects.filter(order=order, product__seller=request.user)

        # Get shipping address for this order
        shipping_address = ShippingAddress.objects.filter(order=order).first()

        context = {
            'order': order,
            'order_items_for_seller': order_items_for_seller,
            'shipping_address': shipping_address,
        }
        return render(request, 'sellers/seller_order_detail.html', context)
    except Order.DoesNotExist:
        messages.error(request, "Order not found or you do not have permission to view it.")
        return redirect('customers:seller_dashboard')
    except Exception as e:
        messages.error(request, "An error occurred while fetching order details.")
        print(f"Error in seller_order_detail_view (Transaction ID: {order_transaction_id}, User: {request.user.username}): Type: {type(e).__name__}, Details: {e}")
        return redirect('customers:seller_dashboard')

@login_required
def customer_order_history_view(request):
    """
    Customer view to display their order history.
    Lists orders that are not in the default "cart" status.
    """
    try:
        from home.models import Order, OrderStatus # Ensure OrderStatus is imported
        
        default_cart_status = OrderStatus.objects.filter(is_default=True).first()
        
        if not default_cart_status:
            messages.error(request, "System configuration error: Default order status not set. Please contact support.")
            print("CRITICAL: No default OrderStatus found in customer_order_history_view.")
            orders = Order.objects.none() # Return an empty queryset to prevent further errors
        else:
        # Fetch orders that are in a "final" state (e.g., Delivered, Cancelled)
        # or any status that isn't the default "cart" status.
            orders = Order.objects.filter(user=request.user).exclude(status=default_cart_status).order_by('-date_order')
        # Alternatively, if you strictly want only 'is_final' statuses:
        # orders = Order.objects.filter(user=request.user, status__is_final=True).order_by('-date_order')
        context = {
            'orders': orders,
        }
        return render(request, 'customers/order_history.html', context)
    except Exception as e:
        messages.error(request, "An error occurred while fetching your order history.")
        print(f"Error in customer_order_history_view: {str(e)}")
        # Redirect to a safe page, like the profile view, using the correct namespace.
        return redirect('customers:view_profile')

@login_required
def customer_order_detail_view(request, order_transaction_id):
    """
    Customer view to display details of a specific order.
    Ensures the order belongs to the current user and is not their active cart.
    """
    try:
        from home.models import Order, OrderItem, ShippingAddress, OrderStatus # Import OrderStatus
        # Fetch the specific order, ensuring it belongs to the current user
        # and is not their active "Pending" cart.
        default_cart_status = OrderStatus.objects.filter(is_default=True).first()
        order = get_object_or_404(Order, transaction_id=order_transaction_id, user=request.user)
        if order.status == default_cart_status:
            raise Order.DoesNotExist # Or handle as "cart view" if you want them to see pending cart here
        # Get all order items for this order
        order_items = OrderItem.objects.filter(order=order)

        # Get shipping address for this order
        shipping_address = ShippingAddress.objects.filter(order=order).first()

        context = {
            'order': order,
            'order_items': order_items,
            'shipping_address': shipping_address,
        }
        return render(request, 'customers/customer_order_detail.html', context) # Correct template path
    except Order.DoesNotExist:
        messages.error(request, "Order not found or you do not have permission to view it.")
        return redirect('customers:customer_order_history') # Correct namespace for redirect
    except Exception as e:
        messages.error(request, "An error occurred while fetching order details.")
        print(f"Error in customer_order_detail_view: {str(e)}")
        return redirect('customers:customer_order_history')
 
@login_required
@staff_member_required
def staff_user_list_view(request):
    """
    Staff view to list all users with their profile information.
    """
    try:
        # Fetch all users and prefetch their related profile to avoid N+1 queries
        # Exclude staff users if you only want to manage customers/sellers here
        # Or include all users if staff can manage other staff too
        users_list = User.objects.select_related('profile__user_level', 'profile__valid_status', 'profile__division').order_by('username')

        paginator = Paginator(users_list, 20) # Show 20 users per page
        page_number = request.GET.get('page')
        users = paginator.get_page(page_number)

        context = {
            'users': users,
        }
        return render(request, 'staff/user_list.html', context)
    except Exception as e:
        messages.error(request, f"An error occurred while fetching the user list: {str(e)}")
        print(f"Error in staff_user_list_view: {str(e)}")
        return redirect('customers:staff_dashboard') # Redirect to dashboard on error

@login_required
@staff_member_required
def staff_user_detail_view(request, pk):
    """
    Staff view to display details of a specific user.
    """
    try:
        # Fetch the user and prefetch related profile data
        user_obj = get_object_or_404(
            User.objects.select_related('profile__user_level', 'profile__valid_status', 'profile__division', 'profile__seller_specific_profile'),
            pk=pk
        )
        context = {
            'user_obj': user_obj,
        }
        return render(request, 'staff/user_detail.html', context)
    except Exception as e:
        messages.error(request, f"An error occurred while fetching user details: {str(e)}")
        print(f"Error in staff_user_detail_view (User PK: {pk}): {str(e)}")
        return redirect('customers:staff_user_list') # Redirect back to user list on error

@login_required
@staff_member_required # General staff can view, permissions can be refined later
def staff_user_orders_view(request, user_id):
    """
    Staff view to list all orders for a specific user.
    """
    target_user = get_object_or_404(User, pk=user_id)
    from home.models import Order # Local import

    order_list = Order.objects.filter(user=target_user).order_by('-date_order')

    paginator = Paginator(order_list, 15) # Show 15 orders per page
    page_number = request.GET.get('page')
    try:
        orders = paginator.page(page_number)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    context = {
        'orders': orders,
        'target_user': target_user,
        'page_title': f"Orders for {target_user.username}",
    }
    # We can reuse the 'staff/all_orders.html' template if we make it flexible
    # or create a new one 'staff/user_orders.html'
    return render(request, 'staff/all_orders.html', context) # Reusing for now

@login_required
@admin_or_hr_required # Restrict editing to SDM or Superusers
def staff_user_edit_view(request, pk):
    """
    Staff view to edit details of a specific user.
    Restricted to SDM division and Superusers.
    """
    try:
        # Fetch the user and related profile data
        user_obj = get_object_or_404(User.objects.select_related('profile'), pk=pk)
        profile = user_obj.profile # Assuming profile always exists due to signal

        # Handle SellerProfile separately if needed, or include fields in StaffProfileEditForm
        # For simplicity now, we'll just edit User and Profile.
        # If you need to edit SellerProfile fields, you'd fetch it here:
        # seller_profile = hasattr(profile, 'seller_specific_profile') ? profile.seller_specific_profile : None

        if request.method == 'POST':
            user_form = StaffUserBaseForm(request.POST, instance=user_obj, prefix="user")
            profile_form = StaffProfileEditForm(request.POST, instance=profile, prefix="profile")

            if user_form.is_valid() and profile_form.is_valid():
                user_form.save()
                profile_form.save()
                # If you had a seller_form, you'd validate and save it here too.

                messages.success(request, f"User '{user_obj.username}' details updated successfully.")
                return redirect('customers:staff_user_detail', pk=user_obj.pk)
            else:
                messages.error(request, "Failed to update user details. Please review the errors below.")
        else:
            user_form = StaffUserBaseForm(instance=user_obj, prefix="user")
            profile_form = StaffProfileEditForm(instance=profile, prefix="profile")
            # If you had a seller_form, instantiate it here:
            # seller_form = StaffSellerProfileEditForm(instance=seller_profile, prefix="seller")

        context = {
            'user_obj': user_obj,
            'user_form': user_form,
            'profile_form': profile_form,
            # 'seller_form': seller_form, # Include if you add seller form
        }
        return render(request, 'staff/user_edit.html', context)
    except Exception as e:
        messages.error(request, f"An error occurred while loading the edit page for user {pk}: {str(e)}")
        print(f"Error in staff_user_edit_view (User PK: {pk}): {str(e)}")
        return redirect('customers:staff_user_list') # Redirect back to user list on error

@login_required
@staff_member_required # Or a more specific permission decorator if needed
def staff_product_list_view(request):
    """
    Staff view to list all products in the system.
    Allows for future management like editing, delisting, etc.
    """
    try:
        products_list = Product.objects.all().select_related('category', 'seller').order_by('-upload_time')

        paginator = Paginator(products_list, 20) # Show 20 products per page
        page_number = request.GET.get('page')
        products = paginator.get_page(page_number)

        context = {
            'products': products,
            'page_title': "Manage All Products"
        }
        return render(request, 'staff/product_list.html', context)
    except Exception as e:
        messages.error(request, f"An error occurred while fetching the product list: {str(e)}")
        print(f"Error in staff_product_list_view: {str(e)}")
        return redirect('customers:staff_dashboard')

@login_required
@staff_member_required # Or a more specific permission like @admin_or_hr_required
def staff_product_delete_view(request, pk):
    """
    Staff view to delete a product.
    """
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product_name = product.name # Get name before deleting for the message
        try:
            product.delete()
            messages.success(request, f"Product '{product_name}' has been deleted successfully by staff.")
        except Exception as e:
            messages.error(request, f"An error occurred while trying to delete product '{product_name}': {str(e)}")
            print(f"Error in staff_product_delete_view (POST) for product PK {pk}: {str(e)}")
        return redirect('customers:staff_product_list')
    
    context = {
        'product': product,
        'page_title': f"Confirm Delete: {product.name}"
    }
    return render(request, 'staff/product_confirm_delete.html', context)

@login_required
@pimpinan_required
def pimpinan_dashboard_view(request):
    """
    Dashboard view for Pimpinan, showing key business analytics.
    This view can serve both the initial page and the API endpoint for chart data.
    """
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        
        # 1. PIE CHART: Penjualan per Kategori
        sales_per_category = OrderItem.objects.values('product__category__name').annotate(
            total_quantity_sold=Sum('quantity')
        ).order_by('-total_quantity_sold')

        pie_chart_data = {
            "labels": [item['product__category__name'] for item in sales_per_category],
            "data": [item['total_quantity_sold'] for item in sales_per_category],
        }

        # 2. PIE CHART: Jumlah Pengguna per Role
        users_per_level = Profile.objects.values('user_level__name').annotate(
            user_count=Count('id')
        ).order_by('-user_count')

        user_role_data = {
            "labels": [item['user_level__name'] for item in users_per_level if item['user_level__name']],
            "data": [item['user_count'] for item in users_per_level if item['user_level__name']],
        }

        # 3. LINE CHART: Penjualan per Bulan
        sales_per_month = OrderItem.objects.annotate(
            month=TruncMonth('order__date_order')
        ).values('month').annotate(
            total_monthly_quantity=Sum('quantity')
        ).order_by('month')
        
        line_chart_data = {
            "labels": [item['month'].strftime('%B %Y') for item in sales_per_month],
            "data": [item['total_monthly_quantity'] for item in sales_per_month],
        }

        # 4. BAR CHART: Produk Terlaris
        best_selling_products = OrderItem.objects.values('product__name').annotate(
            total_sold=Sum('quantity')
        ).order_by('-total_sold')[:5]

        bar_chart_data = {
            "labels": [item['product__name'] for item in best_selling_products],
            "data": [item['total_sold'] for item in best_selling_products],
        }

        # Combine all data into a single response
        all_chart_data = {
            'sales_per_category': pie_chart_data,
            'user_roles': user_role_data,
            'monthly_sales': line_chart_data,
            'top_products': bar_chart_data,
        }
        return JsonResponse(all_chart_data)

    context = { 'page_title': "Dashboard Pimpinan" }
    return render(request, 'staff/pimpinan_dashboard.html', context)

@login_required
@pimpinan_required
def export_pimpinan_dashboard_pdf(request):
    # 1. Get the same data as the dashboard view
    # Sales by Category
    sales_data = OrderItem.objects.values('product__category__name').annotate(total_quantity_sold=Sum('quantity')).order_by('-total_quantity_sold')
    sales_per_category = [{'labels': item['product__category__name'], 'data': item['total_quantity_sold']} for item in sales_data]

    # Monthly Sales
    monthly_data = OrderItem.objects.annotate(month=TruncMonth('order__date_order')).values('month').annotate(total_monthly_quantity=Sum('quantity')).order_by('month')
    monthly_sales = [{'labels': item['month'].strftime('%B %Y'), 'data': item['total_monthly_quantity']} for item in monthly_data]

    # Top Products
    top_data = OrderItem.objects.values('product__name').annotate(total_sold=Sum('quantity')).order_by('-total_sold')[:5]
    top_products = [{'labels': item['product__name'], 'data': item['total_sold']} for item in top_data]

    context = {
        'sales_per_category': sales_per_category,
        'monthly_sales': monthly_sales,
        'top_products': top_products,
    }

    # 2. Render HTML template
    template = get_template('staff/pimpinan_dashboard_report.html')
    html = template.render(context)

    # 3. Create a PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pimpinan_report.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@login_required
@pimpinan_required
def export_pimpinan_dashboard_xlsx(request):
    # 1. Get the data again
    sales_data = OrderItem.objects.values('product__category__name').annotate(total_quantity_sold=Sum('quantity')).order_by('-total_quantity_sold')
    monthly_data = OrderItem.objects.annotate(month=TruncMonth('order__date_order')).values('month').annotate(total_monthly_quantity=Sum('quantity')).order_by('month')
    top_data = OrderItem.objects.values('product__name').annotate(total_sold=Sum('quantity')).order_by('-total_sold')[:5]

    # 2. Create an Excel workbook and sheets
    workbook = openpyxl.Workbook()
    
    # Sheet 1: Sales by Category
    sheet1 = workbook.active
    sheet1.title = "Sales by Category"
    sheet1.append(['Category', 'Items Sold'])
    for item in sales_data:
        sheet1.append([item['product__category__name'], item['total_quantity_sold']])
        
    # Sheet 2: Monthly Sales
    sheet2 = workbook.create_sheet(title="Monthly Sales")
    sheet2.append(['Month', 'Total Items Sold'])
    for item in monthly_data:
        sheet2.append([item['month'].strftime('%B %Y'), item['total_monthly_quantity']])

    # Sheet 3: Top Products
    sheet3 = workbook.create_sheet(title="Top Products")
    sheet3.append(['Product', 'Quantity Sold'])
    for item in top_data:
        sheet3.append([item['product__name'], item['total_sold']])

    # 3. Create the response
    response = HttpResponse(
        content=save_virtual_workbook(workbook),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pimpinan_report.xlsx"'
    return response